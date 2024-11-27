
import tkinter as tk
from tkinter import messagebox
import speech_recognition as sr
from openpyxl import Workbook
import threading

class Spreadsheet(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel-like Interface with Speech Recognition")
        self.geometry("600x400")

        self.recognizing = False
        self.recognizer_thread = None

        self.table = tk.Frame(self)
        self.table.pack(expand=True, fill=tk.BOTH)

        self.headers = ["Customer Name", "Item", "Quantity", "Price"]
        self.cells = {}

        for j, header in enumerate(self.headers):
            label = tk.Label(self.table, text=header, borderwidth=2, relief="ridge")
            label.grid(row=0, column=j, sticky="nsew")
            self.table.grid_columnconfigure(j, weight=1)

        for i in range(1, 11):
            for j in range(len(self.headers)):
                cell = tk.Entry(self.table, width=10)
                cell.grid(row=i, column=j, sticky="nsew")
                self.cells[(i, j)] = cell

        for i in range(1, 11):
            self.table.grid_rowconfigure(i, weight=1)

        self.recognize_button = tk.Button(self, text="Start Speech Recognition", command=self.start_recognition)
        self.recognize_button.pack(pady=10)

        # self.pause_button = tk.Button(self, text="Pause Speech Recognition", command=self.pause_recognition)
        # self.pause_button.pack(pady=10)

        self.save_button = tk.Button(self, text="Save to Excel", command=self.save_to_excel)
        self.save_button.pack(pady=10)

    def start_recognition(self):
        self.recognizing = True
        self.recognizer_thread = threading.Thread(target=self.fill_cells_continuously)
        self.recognizer_thread.start()

    # def pause_recognition(self):
    #     self.recognizing = False
    #     if self.recognizer_thread:
    #         self.recognizer_thread.join()

    def fill_cells_continuously(self):
        while self.recognizing:
            row, col = self.get_selected_cell()
            if row is not None and col is not None:
                text = recognize_speech()
                if text:
                    if text == "next":
                        next_col = (col + 1) % len(self.headers)
                        next_row = row + (col + 1) // len(self.headers)
                        self.cells[(next_row % 11, next_col)].focus_set()
                    elif text == "down":
                        next_row = (row + 1) % 11
                        self.cells[(next_row, col)].focus_set()
                    else:
                        self.cells[(row, col)].insert(0, text)

    def get_selected_cell(self):
        for (row, col), cell in self.cells.items():
            if (cell.focus_get() == cell):
                return row, col
            else:
                return None, None

    def save_to_excel(self):
        workbook = Workbook()
        sheet = workbook.active
        for j, header in enumerate(self.headers):
            sheet.cell(row=1, column=j+1, value=header)
        for (row, col), cell in self.cells.items():
            sheet.cell(row=row+1, column=col+1, value=cell.get())
        workbook.save("spreadsheet.xlsx")
        messagebox.showinfo("Info", "Spreadsheet saved as 'spreadsheet.xlsx'")

def recognize_speech():
    recognizer = sr.Recognizer()
    with sr.Microphone() as source:
        audio = recognizer.listen(source)
    try:
        text = recognizer.recognize_google(audio).lower()
        return text
    except sr.UnknownValueError:
        messagebox.showerror("Error", "Google Speech Recognition could not understand audio")
        return None
    except sr.RequestError as e:
        messagebox.showerror("Error", f"Could not request results; {e}")
        return None

if __name__ == "__main__":
    app = Spreadsheet()
    app.mainloop()
